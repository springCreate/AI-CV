"""
Excel 导出服务

按 PRD 6.1 节导出标准化岗位匹配清单
"""
import os
import logging
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.models.job import Job, JobMatchRecord

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Excel 导出器"""

    # 表头样式
    HEADER_FILL = PatternFill(start_color="2c5282", end_color="2c5282", fill_type="solid")
    HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    BODY_FONT = Font(name="微软雅黑", size=10)
    THIN_BORDER = Border(
        left=Side(style="thin", color="cbd5e0"),
        right=Side(style="thin", color="cbd5e0"),
        top=Side(style="thin", color="cbd5e0"),
        bottom=Side(style="thin", color="cbd5e0"),
    )

    HEADERS = [
        ("序号", 6),
        ("公司名称", 22),
        ("岗位名称", 22),
        ("工作地点", 12),
        ("薪资", 14),
        ("匹配分数", 10),
        ("工作年限", 10),
        ("学历", 8),
        ("岗位要求摘要", 50),
        ("发布时间", 18),
        ("投递状态", 10),
        ("岗位链接", 30),
        ("平台", 10),
    ]

    @classmethod
    def export_match_list(cls, records: List[JobMatchRecord], filename: str = None) -> str:
        """
        导出匹配清单到 Excel

        Args:
            records: 匹配记录列表（含关联 job）
            filename: 文件名（不含路径）

        Returns:
            Excel 文件路径
        """
        from flask import current_app
        storage_cfg = current_app.config["STORAGE_CFG"]
        out_dir = Path(current_app.config["BASE_DIR"]) / storage_cfg["export_dir"]
        out_dir.mkdir(parents=True, exist_ok=True)

        if not filename:
            filename = f"jobs_export_{len(records)}.xlsx"
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"
        # 文件名安全处理
        filename = "".join(c if c.isalnum() or c in "._-" else "_" for c in filename)
        file_path = out_dir / filename

        wb = Workbook()
        ws = wb.active
        ws.title = "岗位匹配清单"

        # 写表头
        for col_idx, (title, width) in enumerate(cls.HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = cls.THIN_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 30

        # 写数据
        # 按匹配分数倒序
        sorted_records = sorted(
            [r for r in records if r.hard_filter_passed],
            key=lambda r: r.match_score or 0,
            reverse=True,
        )

        for idx, rec in enumerate(sorted_records, 1):
            job = rec.job
            if not job:
                continue

            # 投递状态
            from app.models.application import ApplicationRecord
            app_record = ApplicationRecord.query.filter_by(
                user_id=rec.user_id, job_id=job.id
            ).first()
            status_text = {
                "not_applied": "未投递",
                "applied": "已投递",
                "interview": "面试中",
                "offer": "Offer",
                "rejected": "已拒绝",
            }.get(app_record.status if app_record else "not_applied", "未投递")

            # JD 摘要（前 200 字）
            jd_summary = (job.jd_text or "")[:200]
            if len(job.jd_text or "") > 200:
                jd_summary += "..."

            # 发布时间格式化
            publish_time = ""
            if job.publish_time:
                try:
                    publish_time = job.publish_time.strftime("%Y-%m-%d %H:%M")
                except Exception:
                    publish_time = str(job.publish_time)

            row_data = [
                idx,
                job.company or "",
                job.title or "",
                f"{job.city or ''} {job.district or ''}".strip(),
                job.salary_text or f"{job.salary_min or ''}-{job.salary_max or ''}",
                rec.match_score or 0,
                job.work_years or "不限",
                job.education or "不限",
                jd_summary,
                publish_time,
                status_text,
                job.job_url or "",
                job.platform or "",
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=idx + 1, column=col_idx, value=value)
                cell.font = cls.BODY_FONT
                cell.alignment = Alignment(vertical="center", wrap_text=True,
                                           horizontal="center" if col_idx in (1, 6, 11) else "left")
                cell.border = cls.THIN_BORDER

                # 匹配分数着色
                if col_idx == 6:
                    score = rec.match_score or 0
                    if score >= 80:
                        cell.fill = PatternFill(start_color="9ae6b4", end_color="9ae6b4", fill_type="solid")
                    elif score >= 60:
                        cell.fill = PatternFill(start_color="fefcbf", end_color="fefcbf", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="fed7d7", end_color="fed7d7", fill_type="solid")

            ws.row_dimensions[idx + 1].height = 60

        # 冻结首行
        ws.freeze_panes = "A2"

        # 添加统计 sheet
        ws2 = wb.create_sheet("统计汇总")
        stats = [
            ("总匹配数", len(sorted_records)),
            ("高匹配度（≥80）", sum(1 for r in sorted_records if (r.match_score or 0) >= 80)),
            ("中匹配度（60-79）", sum(1 for r in sorted_records if 60 <= (r.match_score or 0) < 80)),
            ("低匹配度（<60）", sum(1 for r in sorted_records if (r.match_score or 0) < 60)),
            ("已投递", sum(1 for r in sorted_records if ApplicationRecord.query.filter_by(
                user_id=r.user_id, job_id=r.job_id).first())),
        ]
        for i, (label, value) in enumerate(stats, 1):
            ws2.cell(row=i, column=1, value=label).font = Font(name="微软雅黑", size=11, bold=True)
            ws2.cell(row=i, column=2, value=value).font = Font(name="微软雅黑", size=11)
        ws2.column_dimensions["A"].width = 22
        ws2.column_dimensions["B"].width = 12

        wb.save(str(file_path))
        logger.info("Excel 已导出: %s（%d 条记录）", file_path, len(sorted_records))
        return str(file_path)
